import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from kinetics.internal.models.solution_data import SolutionData


def create_excel_solution(solution: SolutionData):
    wb = Workbook()
    ws = wb.active

    title_font = Font(name="Times New Roman", size=12, bold=True)

    # заголовок и текущее время
    report_title = "Отчёт по расчету прямой задачи химической кинетики"
    report_time = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    ws['A1'] = f'{report_title} ({report_time})'
    ws['A1'].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.row_dimensions[1].height = 30

    # подзаголовок "Входные данные"
    input_data_title = "Входные данные"
    ws['A2'] = input_data_title
    ws['A2'].font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.row_dimensions[2].height = 20

    input_data_headers = ['Начальное время t0', 'Время t', 'Шаг','Метод решения', 'Матрица стехиометрических коэффициентов',
                          'Матрица показателей степени', 'Экспериментальные данные', 'Константы скорости']

    input_data_values = [solution.input_data.initial_time, solution.input_data.time, solution.input_data.step,
                         solution.input_data.method,
                         solution.input_data.matrix_stechiometric_coefficients,
                         solution.input_data.matrix_indicators, solution.input_data.experimental_data,
                         solution.input_data.constants_speed]

    for i, header in enumerate(input_data_headers):
        ws.cell(row=3, column=i+1).value = header
        ws.cell(row=3, column=i+1).font = ws.cell(row=3, column=i+1).font.copy(bold=True)
        ws.cell(row=3, column=i+1).border = ws.cell(row=3, column=i+1).border.copy(bottom=None)
        ws.cell(row=3, column=i+1).alignment = ws.cell(row=3, column=i+1).alignment.copy(horizontal='center')
        ws.column_dimensions[get_column_letter(i+1)].width = 20

        value = input_data_values[i]
        if isinstance(value, list):
            for j, v in enumerate(value):
                ws.cell(row=4+j, column=i+1).value = v
        else:
            ws.cell(row=4, column=i+1).value = value

    # подзаголовок "Результаты расчета"
    input_data_title = "Результаты расчета"
    ws['A5'] = input_data_title
    ws['A5'].font = title_font
    ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=4)
    ws.row_dimensions[5].height = 20

    result_headers = ['Результаты', 'Время t', 'Значения в эксперементальных точках']

    result_values = [solution.result, solution.time, solution.experimental_point]

    for i, header in enumerate(result_headers):
        ws.cell(row=7, column=i+1).value = header
        ws.cell(row=7, column=i+1).font = ws.cell(row=7, column=i+1).font.copy(bold=True)
        ws.cell(row=7, column=i+1).border = ws.cell(row=7, column=i+1).border.copy(bottom=None)
        ws.cell(row=7, column=i+1).alignment = ws.cell(row=7, column=i+1).alignment.copy(horizontal='center')
        ws.column_dimensions[get_column_letter(i+1)].width = 20

        value = result_values[i]
        if isinstance(value, list):
            for j, v in enumerate(value):
                ws.cell(row=8+j, column=i+1).value = v
        else:
            ws.cell(row=8, column=i+1).value = value

    # Сохраните workbook
    wb.save(f'Report({report_time}).xlsx')
